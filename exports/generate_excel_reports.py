"""
Generate the five formatted Excel deliverables into exports/excel/.

Run:  python exports/generate_excel_reports.py

Data is pulled from the live project modules, so the workbooks cannot drift from
the running system. Requires pandas and openpyxl.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# The Windows console defaults to cp1252, which cannot encode the project path
# or the bilingual labels. Reconfigure rather than let a print() kill the run
# after the workbooks are already on disk.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, OSError):
    pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import constants as C          # noqa: E402
import engine as E             # noqa: E402

OUT = ROOT / "exports" / "excel"
OUT.mkdir(parents=True, exist_ok=True)

# --- shared styling -------------------------------------------------------
NAVY, FOREST, SLATE, OCEAN, PLUM = "1B365D", "1E4D2B", "2C3E50", "005A9C", "4A2E35"
ZEBRA = PatternFill("solid", fgColor="F4F6F8")
AMBER = PatternFill("solid", fgColor="FFE0B2")
RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
THIN = Side(style="thin", color="D0D5DA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, header_hex: str, n_cols: int, freeze: str = "A2",
                zebra: bool = True, width: int = 20) -> None:
    """Header fill + white bold text, borders, zebra striping, freeze panes."""
    fill = PatternFill("solid", fgColor=header_hex)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze
    for r in range(2, ws.max_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if zebra and r % 2 == 0:
                cell.fill = ZEBRA


def write_df(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col))
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            ws.cell(row=i, column=j,
                    value=(None if pd.isna(v) else
                           (float(v) if isinstance(v, (int, float)) else str(v))))


def note(ws, row: int, text: str) -> None:
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(italic=True, size=9, color="666666")
    c.alignment = Alignment(wrap_text=True, vertical="top")


# ==========================================================================
# 1. Module 1 — acid dosing
# ==========================================================================

def build_module1() -> Path:
    hno3 = C.FERTILISERS["hno3_38"]
    h3po4 = C.FERTILISERS["h3po4_59"]
    m_n = E.acid_molarity_mol_per_l(hno3)
    m_p = E.acid_molarity_mol_per_l(h3po4)
    buf = C.DEFAULT_POLICY.hco3_buffer_mmol_l

    rows = []
    for i in range(0, 15):
        hco3 = 1.0 + i * 0.5
        h = max(0.0, hco3 - buf)
        ml_m3 = h / m_n * 1000.0
        rows.append({
            "Raw HCO3 (mmol/L) 原水碳酸氢盐": round(hco3, 2),
            "HCO3 Buffer Retained (mmol/L) 保留缓冲": buf,
            "H+ Required (mmol/L) 所需质子": round(h, 3),
            "HNO3 38% (mL/m3) 硝酸": round(ml_m3, 1),
            "HNO3 38% (L per 1000L stock @100x) 母液罐": round(h * 167 * 0.1 / 1.24, 2),
            "H3PO4 59% (mL/m3) 磷酸": round(h / m_p * 1000.0, 1),
            "N-NO3 added (mmol/L) 带入硝态氮": round(h, 3),
            "P added if phosphoric (mmol/L) 带入磷": round(h, 3),
        })
    df1 = pd.DataFrame(rows)

    df2 = pd.DataFrame([
        {"Acid 酸": f.name_en, "Chinese 中文": f.name_zh,
         "Density (kg/L) 密度": f.density,
         "g product / mol H+ 每摩尔质子克数": f.mass_per_mol_ion,
         "Molarity (mol H+/L) 摩尔浓度": round(E.acid_molarity_mol_per_l(f), 4),
         "Anion per mol H+ 伴随阴离子": "NO3-" if "hno3" in f.fid else "H2PO4-",
         "mL per (1 mmol/L x m3) 单位用量": round(1000.0 / E.acid_molarity_mol_per_l(f), 3)}
        for f in (C.FERTILISERS["hno3_38"], C.FERTILISERS["hno3_60"], h3po4)])

    path = OUT / "Module1_Acid_Dosing_Dashboard.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df1.to_excel(xl, sheet_name="Acid Titration Lookup", index=False)
        df2.to_excel(xl, sheet_name="Nutrient Yield Breakdown", index=False)
        wb = xl.book

        ws = wb["Acid Titration Lookup"]
        style_sheet(ws, NAVY, len(df1.columns), width=22)
        # amber highlight where raw HCO3 exceeds 4.0 mmol/L
        ws.conditional_formatting.add(
            f"A2:A{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["4.0"], fill=AMBER))
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            DataBarRule(start_type="min", end_type="max", color=NAVY))

        ch = LineChart()
        ch.title = "Raw Water HCO3 vs Acid Required (mL/m3)"
        ch.y_axis.title = "HNO3 38% required (mL/m3)"
        ch.x_axis.title = "Raw water HCO3 (mmol/L)"
        ch.height, ch.width = 9, 18
        ch.add_data(Reference(ws, min_col=4, min_row=1, max_row=ws.max_row),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(ch, f"A{ws.max_row + 3}")
        note(ws, ws.max_row + 2,
             f"Buffer of {buf} mmol/L HCO3 is retained deliberately (WUR p.24): "
             "neutralising all of it drops irrigation pH below 5. "
             "保留碳酸氢盐缓冲以将 pH 稳定在 5.5-6.0。")

        ws2 = wb["Nutrient Yield Breakdown"]
        style_sheet(ws2, NAVY, len(df2.columns), width=24)
        note(ws2, ws2.max_row + 2,
             "DIV-1: the export brief specifies 65% HNO3 / 85% H3PO4; WUR "
             "Table 5 (p.26) publishes 38% and 60% nitric and 59% phosphoric. "
             "Catalogue grades are shown. 说明书与手册不一致，此处采用手册数值。")
    return path


# ==========================================================================
# 2. Module 2 — crop target matrix
# ==========================================================================

def build_module2() -> Path:
    df = pd.read_csv(ROOT / "exports" / "data" / "wur_crop_targets.csv")
    path = OUT / "Module2_WUR_Crop_Target_Matrix.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        for cat in C.CROP_CATEGORIES:
            sub = df[df["category"] == cat]
            if sub.empty:
                continue
            name = cat.replace("_", " ").title()[:31]
            sub.to_excel(xl, sheet_name=name, index=False)
            ws = xl.book[name]
            style_sheet(ws, FOREST, len(sub.columns), freeze="D2", width=16)

            ec_col = list(sub.columns).index("ec_fertigation") + 1
            ws.conditional_formatting.add(
                f"{get_column_letter(ec_col)}2:{get_column_letter(ec_col)}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="FFFFFF",
                               end_type="max", end_color=FOREST))

            ch = BarChart()
            ch.type, ch.style = "col", 10
            ch.title = f"{name}: Target EC and key cations by crop/stage"
            ch.y_axis.title = "mmol/L  ·  EC in mS/cm"
            ch.height, ch.width = 10, 26
            cols = {c: list(sub.columns).index(c) + 1
                    for c in ("ec_fertigation", "K_mmol_l", "Ca_mmol_l", "Mg_mmol_l")
                    if c in sub.columns}
            last = min(ws.max_row, 26)          # keep the chart legible
            for ci in cols.values():
                ch.add_data(Reference(ws, min_col=ci, min_row=1, max_row=last),
                            titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=2, min_row=2, max_row=last))
            ws.add_chart(ch, f"A{ws.max_row + 3}")
    return path


# ==========================================================================
# 3. Module 3 — diagnostics & safety gates
# ==========================================================================

def build_module3() -> Path:
    crop = C.get_crop("tomato")
    sample = {"NH4": 0.1, "K": 7.2, "Na": 2.7, "Ca": 11.4, "Mg": 6.0,
              "NO3": 22.1, "Cl": 2.8, "S": 7.9, "P": 2.8, "HCO3": 0.4}
    cat = E.eq_cations(sample)
    an = E.eq_anions(sample)
    err = abs(cat - an) / max(cat, an) * 100

    rows = []
    for ion, meas in sample.items():
        tgt = crop.root_zone_targets.get(ion)
        rows.append({
            "Ion 离子": ion,
            "Measured (mmol/L) 实测": meas,
            "Target (mmol/L) 目标": tgt if tgt is not None else None,
            "Deviation % 偏差": (round((meas - tgt) / tgt * 100, 1)
                               if tgt else None),
            "Charge 电荷": C.ION_CHARGE.get(ion, 1),
            "mmol_c/L 当量": round(C.ION_CHARGE.get(ion, 1) * meas, 3),
            "Side 侧": "cation" if ion in C.CATIONS else "anion",
        })
    rows.append({"Ion 离子": "-- Eq cations --", "mmol_c/L 当量": round(cat, 3)})
    rows.append({"Ion 离子": "-- Eq anions --", "mmol_c/L 当量": round(an, 3)})
    rows.append({"Ion 离子": "-- Charge error % --", "mmol_c/L 当量": round(err, 2)})
    rows.append({"Ion 离子": "-- Estimated EC --",
                 "mmol_c/L 当量": round(E.ec_from_ions(sample), 3)})
    df1 = pd.DataFrame(rows)

    import json
    reg = json.loads((ROOT / "exports" / "data" /
                      "module3_safety_gates_and_rules.json")
                     .read_text(encoding="utf-8"))["gate_registry"]
    df2 = pd.DataFrame([{"Gate ID 闸门": g["gate_id"], "Severity 级别": g["severity"],
                         "Module 模块": g["module"], "Condition 触发条件": g["condition"],
                         "Action 处置": g["action"], "Provenance 来源": g["provenance"]}
                        for g in reg])

    path = OUT / "Module3_Diagnostic_Safety_Simulator.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df1.to_excel(xl, sheet_name="Diagnostic Evaluator", index=False)
        df2.to_excel(xl, sheet_name="Safety Gate Thresholds", index=False)
        wb = xl.book

        ws = wb["Diagnostic Evaluator"]
        style_sheet(ws, SLATE, len(df1.columns), width=20)
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["50"], fill=RED))
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["-50"], fill=RED))
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            CellIsRule(operator="between", formula=["25", "50"], fill=AMBER))

        ch = BarChart()
        ch.type, ch.style = "col", 10
        ch.title = "Measured vs target root-zone concentration (mmol/L)"
        ch.height, ch.width = 10, 20
        last = len(sample) + 1
        ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=last),
                    titles_from_data=True)
        ch.add_data(Reference(ws, min_col=3, min_row=1, max_row=last),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(ch, f"A{ws.max_row + 3}")
        note(ws, ws.max_row + 2,
             "Charge balance includes H+ from acid dosing (DIV-4). Sample is the "
             "Eurofins Optifeed analysis reproduced on WUR p.29. "
             "电荷平衡含加酸带入的 H+。")

        ws2 = wb["Safety Gate Thresholds"]
        style_sheet(ws2, SLATE, len(df2.columns), width=26)
        ws2.conditional_formatting.add(
            f"B2:B{ws2.max_row}",
            CellIsRule(operator="equal", formula=['"BLOCKING"'], fill=RED))
        ws2.conditional_formatting.add(
            f"B2:B{ws2.max_row}",
            CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=AMBER))
        ws2.conditional_formatting.add(
            f"B2:B{ws2.max_row}",
            CellIsRule(operator="equal", formula=['"INFO"'], fill=GREEN))
        note(ws2, ws2.max_row + 2,
             "DIV-2/DIV-3: the brief's G-ACID-POISON and G-SALINITY-MELTDOWN are "
             "one merged gate (G-MELTDOWN); G-K-CA-ANTAGONISM (dK>2.0) is NOT "
             "implemented and was not fabricated. 说明书部分闸门名称与实现不一致。")
    return path


# ==========================================================================
# 4. Module 4 — leaching fraction model
# ==========================================================================

def build_module4() -> Path:
    v_irr = 5.0
    rows = []
    for lf in [x / 2 for x in range(10, 101, 5)]:      # 5% .. 50%
        drain = v_irr * lf / 100.0
        r = E.evaluate_leaching(v_irr, drain, 2.0, 4.0)
        rows.append({
            "Current LF % 当前排液比": round(lf, 1),
            "Drain (L/m2) 排液量": round(drain, 3),
            "Uptake (L/m2) 吸水量": round(r.uptake_l_m2, 3),
            "Wash Case 冲洗分级": r.wash_case,
            "Target LF % 目标排液比": round(r.target_lf_pct, 1),
            "Extra Irrigation (L/m2/day) 需增加灌溉量":
                round(r.extra_irrigation_l_m2, 3),
            "Target Irrigation (L/m2/day) 目标灌溉量":
                round(r.target_irrigation_l_m2, 3),
            "Anomaly 异常": "YES" if r.is_wash_anomaly else "",
        })
    df1 = pd.DataFrame(rows)

    grid = []
    for lf in (10, 15, 20, 25, 30, 35, 40, 45):
        row = {"Current LF % 当前排液比": lf}
        for gap in (0.5, 1.0, 1.5, 2.0, 2.5, 3.0):
            r = E.evaluate_leaching(v_irr, v_irr * lf / 100.0, 2.0, 2.0 + gap)
            row[f"dEC {gap}"] = round(r.extra_irrigation_l_m2, 3)
        grid.append(row)
    df2 = pd.DataFrame(grid)

    path = OUT / "Module4_Leaching_Fraction_Model.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df1.to_excel(xl, sheet_name="LF Simulation", index=False)
        df2.to_excel(xl, sheet_name="LF x EC Gap Matrix", index=False)
        wb = xl.book

        ws = wb["LF Simulation"]
        style_sheet(ws, OCEAN, len(df1.columns), width=22)
        ws.conditional_formatting.add(
            f"F2:F{ws.max_row}",
            DataBarRule(start_type="min", end_type="max", color=OCEAN))
        ws.conditional_formatting.add(
            f"H2:H{ws.max_row}",
            CellIsRule(operator="equal", formula=['"YES"'], fill=RED))

        ch = LineChart()
        ch.title = "Extra irrigation needed as current LF drops (V_irr = 5.0 L/m2)"
        ch.y_axis.title = "Extra irrigation (L/m2/day)"
        ch.x_axis.title = "Current leaching fraction (%)"
        ch.height, ch.width = 10, 22
        ch.add_data(Reference(ws, min_col=6, min_row=1, max_row=ws.max_row),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(ch, f"A{ws.max_row + 3}")
        note(ws, ws.max_row + 2,
             "Tiered target (DIV-6): 32.5% while LF<30%; LF+10 capped at 50% for "
             "30-40%; NO volume target at LF>=40% where more water is the wrong "
             "remedy. 分级目标：40% 以上不再增加水量，改查偏流与电导标定。")

        ws2 = wb["LF x EC Gap Matrix"]
        style_sheet(ws2, OCEAN, len(df2.columns), width=18)
        ws2.conditional_formatting.add(
            f"B2:G{ws2.max_row}",
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color=OCEAN))
        note(ws2, ws2.max_row + 2,
             "Zero below dEC 2.0 = no wash triggered. Zero at LF>=40% = anomaly, "
             "not 'nothing to do'. 电导差低于 2.0 不触发冲洗；排液比 40% 以上为异常。")
    return path


# ==========================================================================
# 5. Module 5 — A/B stock tank calculator
# ==========================================================================

def build_module5() -> Path:
    crop = C.get_crop("tomato")
    steer = E.apply_stage_adjustments(crop, [])
    doses, residual = E.allocate_fertilisers(dict(steer.macro_after),
                                             dict(steer.micro_after))
    split = E.split_ab_tanks(doses)

    rows = []
    for tank, items in (("A", split.tank_a), ("B", split.tank_b)):
        for d in items:
            rows.append({
                "Tank 母液罐": tank,
                "Fertiliser 肥料": d.fert.name_en,
                "Chinese 中文": d.fert.name_zh,
                "Formula 化学式": d.fert.formula,
                "Amount (mmol/L or umol/L) 浓度": round(d.amount_mmol_l, 4),
                "Mass (kg) 质量": round(d.mass_kg, 3),
                "Volume (L) 体积": (round(d.volume_l, 3)
                                  if d.volume_l is not None else None),
                "Micronutrient 微量": "yes" if d.is_micro else "",
            })
    df1 = pd.DataFrame(rows)

    df2 = pd.DataFrame([
        {"Chelate 螯合物": k, "pH min 下限": v[0], "pH max 上限": v[1],
         "Recommended above pH 6.5 高pH推荐":
             "YES" if v[1] >= 10 else "no"}
        for k, v in C.FE_CHELATE_BANDS.items()])

    summary = pd.DataFrame([
        {"Tank 母液罐": "A", "Total mass (kg) 总质量": round(split.mass_a_kg, 2)},
        {"Tank 母液罐": "B", "Total mass (kg) 总质量": round(split.mass_b_kg, 2)}])

    path = OUT / "Module5_AB_Stock_Tank_Calculator.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df1.to_excel(xl, sheet_name="Tank Allocation Breakdown", index=False)
        df2.to_excel(xl, sheet_name="Fe-Chelate pH Selector", index=False)
        wb = xl.book

        ws = wb["Tank Allocation Breakdown"]
        style_sheet(ws, PLUM, len(df1.columns), width=22)
        ws.conditional_formatting.add(
            f"F2:F{ws.max_row}",
            DataBarRule(start_type="min", end_type="max", color=PLUM))

        srow = ws.max_row + 3
        ws.cell(row=srow, column=1, value="Tank 母液罐").font = Font(bold=True)
        ws.cell(row=srow, column=2, value="Total mass (kg) 总质量").font = Font(bold=True)
        for i, (_, r) in enumerate(summary.iterrows(), start=1):
            ws.cell(row=srow + i, column=1, value=r["Tank 母液罐"])
            ws.cell(row=srow + i, column=2, value=float(r["Total mass (kg) 总质量"]))

        ch = DoughnutChart()
        ch.title = "Mass ratio: Tank A vs Tank B (1000 L @ 100x)"
        ch.height, ch.width = 9, 12
        ch.add_data(Reference(ws, min_col=2, min_row=srow, max_row=srow + 2),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=srow + 1,
                                    max_row=srow + 2))
        ws.add_chart(ch, f"D{srow}")
        note(ws, srow + 4,
             "Calcium fertilisers are separated from phosphate and sulphate "
             "absolutely (WUR p.31): at 100x both CaSO4 and Ca3(PO4)2 are far "
             "past saturation. 钙肥必须与磷酸盐、硫酸盐分罐。")

        ws2 = wb["Fe-Chelate pH Selector"]
        style_sheet(ws2, PLUM, len(df2.columns), width=24)
        ws2.conditional_formatting.add(
            f"D2:D{ws2.max_row}",
            CellIsRule(operator="equal", formula=['"YES"'], fill=GREEN))

        ch2 = BarChart()
        ch2.type, ch2.style = "bar", 10
        ch2.title = "Fe-chelate pH stability envelopes (WUR Figure 3a, p.35)"
        ch2.x_axis.title = "pH"
        ch2.height, ch2.width = 8, 16
        ch2.add_data(Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row),
                     titles_from_data=True)
        ch2.add_data(Reference(ws2, min_col=3, min_row=1, max_row=ws2.max_row),
                     titles_from_data=True)
        ch2.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row))
        ws2.add_chart(ch2, f"F2")
        note(ws2, ws2.max_row + 2,
             f"Switch point pH {C.FE_CHELATE_SWITCH_PH} (WUR p.36), not 7.0. "
             "Below it DTPA suffices; above it EDDHA/HBED is strongly "
             "recommended. 切换点为 pH 6.5。")
    return path


# ==========================================================================

def main() -> None:
    builders = [
        ("Module 1 - Acid Dosing Dashboard", build_module1),
        ("Module 2 - WUR Crop Target Matrix", build_module2),
        ("Module 3 - Diagnostic Safety Simulator", build_module3),
        ("Module 4 - Leaching Fraction Model", build_module4),
        ("Module 5 - A/B Stock Tank Calculator", build_module5),
    ]
    print("Generating Excel deliverables into exports/excel/\n")
    ok = 0
    for label, fn in builders:
        try:
            p = fn()
            print(f"  [OK]   {label}\n         {p.name}  "
                  f"({p.stat().st_size:,} bytes)")
            ok += 1
        except Exception as exc:                      # noqa: BLE001
            print(f"  [FAIL] {label}: {type(exc).__name__}: {exc}")
    print(f"\n{ok}/{len(builders)} workbooks written to {OUT}")
    if ok != len(builders):
        sys.exit(1)


if __name__ == "__main__":
    main()
